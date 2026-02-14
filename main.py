import json
import os
import time
import shutil
import pandas as pd
import warnings
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

# ğŸ”§ å±è”½æ— å…³è­¦å‘Š
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# å¢åŠ äº† DESC_FULL æ˜ å°„
COLUMN_MAP = {
    "MARCA": "MARCA",
    "TIPO": "TIPO",
    "PRODUCTO": "PRODUCTO",
    "DESC_FULL": "DESC_FULL", # æ–°å¢è¯¦æƒ…å­—æ®µ
    "Precio ( USD )": "Precio ( USD )",
    "Imagen_Path": "Imagen_Path",
}

WATCH_FOLDER = os.path.dirname(os.path.abspath(__file__))
SOURCES_FILE = os.path.join(WATCH_FOLDER, "sources.json")

def extract_images_for_excel(xlsx_path: str):
    name = os.path.basename(xlsx_path)
    base = os.path.splitext(name)[0]
    img_folder = f"{base}_images"
    img_folder_path = os.path.join(WATCH_FOLDER, img_folder)
    os.makedirs(img_folder_path, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=True)
    sheet = wb.active

    img_col_idx = None
    for col in range(1, sheet.max_column + 1):
        v = sheet.cell(row=1, column=col).value
        if v and str(v).strip().lower() == "imagen":
            img_col_idx = col
            break

    if img_col_idx is None:
        return {}

    image_loader = SheetImageLoader(sheet)
    row_to_path = {}
    img_count = 0

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=img_col_idx)
        addr = cell.coordinate
        if image_loader.image_in(addr):
            try:
                image = image_loader.get(addr)
                img_count += 1
                filename = f"{base}_{img_count:03d}.png"
                out_path = os.path.join(img_folder_path, filename)
                image.save(out_path)
                rel_path = f"{img_folder}/{filename}"
                row_to_path[row] = rel_path
                print(f"ğŸ“¸ ä» {addr} å¯¼å‡ºå›¾ç‰‡ -> {out_path}")
            except:
                pass
    return row_to_path

def excel_to_json(path):
    name = os.path.basename(path)
    if name.startswith("~$") or not name.lower().endswith(".xlsx"):
        return None

    print(f"\nğŸš€ å¼€å§‹å¤„ç†è¡¨æ ¼: {name}")

    try:
        row_to_img = extract_images_for_excel(path)
        df = pd.read_excel(path)

        # âœ¨ æ ¸å¿ƒæ”¹è¿›ï¼šæ‹†åˆ†ã€äº§å“åã€‘å’Œã€è¯¦ç»†ä»‹ç»ã€‘
        titles = []
        descriptions = []
        
        for val in df["PRODUCTO"]:
            raw_text = str(val).strip()
            # æŒ‰ç…§æ¢è¡Œç¬¦æ‹†åˆ†å†…å®¹
            parts = raw_text.split('\n', 1) 
            
            # ç¬¬ä¸€è¡Œæ˜¯çº¢è‰²æ ‡é¢˜
            titles.append(parts[0].strip())
            # å‰©ä¸‹çš„æ‰€æœ‰å†…å®¹æ˜¯è¯¦æƒ…
            descriptions.append(parts[1].strip() if len(parts) > 1 else "")

        df["PRODUCTO"] = titles
        df["DESC_FULL"] = descriptions

        # ä»·æ ¼å…¼å®¹å¤„ç†
        if "Precio ( USD )" not in df.columns:
            for col in df.columns:
                if "precio" in col.lower() and "usd" in col.lower():
                    df["Precio ( USD )"] = df[col]
                    break
        
        if "Precio ( USD )" in df.columns:
            df["Precio ( USD )"] = pd.to_numeric(df["Precio ( USD )"].astype(str).str.replace(r'[^0-9.]', '', regex=True), errors='coerce').fillna(0)

        def is_invalid(val):
            s = str(val).strip()
            return s.startswith('=') or "DISPIMG" in s or s.lower() == "nan" or s == ""

        if "MARCA" in df.columns:
            df = df[~df["MARCA"].apply(is_invalid)]
        
        imagen_paths = []
        for idx in df.index:
            excel_row_num = idx + 2
            imagen_paths.append(row_to_img.get(excel_row_num, ""))
        df["Imagen_Path"] = imagen_paths

        df = df[[c for c in COLUMN_MAP.keys() if c in df.columns]]
        df = df.rename(columns=COLUMN_MAP)

        out_name = os.path.splitext(name)[0] + ".json"
        out_path = os.path.join(WATCH_FOLDER, out_name)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(df.to_dict(orient="records"), f, ensure_ascii=False, indent=2)
        
        print(f"âœ… å¤„ç†å®Œæˆ: {out_name} (å…±ç”Ÿæˆ {len(df)} æ¡æ•°æ®)")
        return out_name
    except Exception as e:
        print(f"âŒ å‘ç”Ÿé”™è¯¯: {e}")
        return None

def build_sources_manifest():
    entries = []
    products_entry = None
    master_json = os.path.join(WATCH_FOLDER, "data.json")
    if os.path.exists(master_json):
        entries.append({"key": "stock", "label": "STOCK", "desc": "CatÃ¡logo de stock actual", "file": "data.json"})

    for name in os.listdir(WATCH_FOLDER):
        if not name.lower().endswith(".json") or name in ("sources.json", "data.json"):
            continue
        base = os.path.splitext(name)[0]
        if name.lower() == "products.json":
            products_entry = {"key": "lista_actual", "label": "LISTA ACTUAL", "desc": "Lista actual", "file": "products.json"}
            continue
        entries.append({"key": base.lower(), "label": base.upper(), "desc": base, "file": name})

    if products_entry: entries.insert(0, products_entry)
    with open(SOURCES_FILE, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)

class Handler(FileSystemEventHandler):
    def on_modified(self, event):
        if not event.is_directory and event.src_path.endswith(".xlsx"):
            if excel_to_json(event.src_path): build_sources_manifest()

if __name__ == "__main__":
    print("=========================================")
    print("ğŸŒŸ STARPHONE æŠ¥è¡¨ç³»ç»Ÿå¯åŠ¨ä¸­...")
    print("=========================================")
    for name in os.listdir(WATCH_FOLDER):
        if name.lower().endswith(".xlsx") and not name.startswith("~$"):
            excel_to_json(os.path.join(WATCH_FOLDER, name))
    build_sources_manifest()
    
    print("\nâœ¨ åˆå§‹åŒ–æ‰«æç»“æŸã€‚æ­£åœ¨ç›‘å¬æ–°æ–‡ä»¶æ”¹åŠ¨...")
    event_handler = Handler()
    observer = Observer()
    observer.schedule(event_handler, WATCH_FOLDER, recursive=False)
    observer.start()
    try:
        while True: time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()